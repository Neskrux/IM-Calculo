import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
def load(p): return json.load(open(os.path.join(ROOT, p), encoding='utf8'))
b9 = load('docs/rodadas/b9/b9-duplicatas-comissao.json')
b10 = load('docs/rodadas/b10/b10-prosoluto-divergente.json')

# 25 distratos (snapshot SELECT 2026-06-02): contrato, unidade, venda_id, data, pro_soluto, valor_venda
distratos = [
 ('25','508 A','11728f4d-67d3-4629-a6b2-d93ec06dbea0','2025-09-23',46482.60,232413.12),
 ('42','609 A','5cfdabb1-25aa-4633-a6e0-b3c4651cf052','2026-02-06',80471.46,402356.42),
 ('79','904 A','39d7aad3-aa7b-459c-820a-17a00e1aa7f4','2026-02-24',83668.40,418341.97),
 ('141','1105 A','c2c465e1-cec4-42bc-8f40-db0d7d7d3106','2026-02-02',85350.13,426750.63),
 ('145','1103 A','472d7dfe-e95a-40d7-8739-328d2b576c44','2026-01-06',74593.99,422991.12),
 ('149','1304 A','e1c0ca9b-452f-4955-aeb3-a0e1fc36642c','2026-02-04',87065.40,435328.32),
 ('168','1702 A','64aa2405-0068-4c20-8d15-736e1a1d6b1b','2025-11-17',88913.16,444567.92),
 ('181','603 B','eb4ac0bc-7de5-49fd-a933-3ab7de549dac','2026-02-02',80470.89,402356.42),
 ('186','803 B','1ed61c21-97b9-4409-bf45-403cb41e3889','2026-01-14',82088.55,410443.77),
 ('209','403 C','8b74adf9-e6dc-404b-8144-d026dbb256ed','2026-01-13',78141.60,375638.57),
 ('255','1405 C','84b74c27-fa97-4cb0-9e22-be95f88237d2','2025-12-29',73581.34,363369.59),
 ('268','604 D','335e0f43-2a77-4399-ae27-eb4af0a56a68','2026-01-06',76639.32,383196.59),
 ('278','901 D','ac19aefe-53c9-4f4c-a519-e603fc6aa05e','2025-12-15',69146.67,345733.70),
 ('284','1001 D','64ffc2fb-b265-41d3-b7d5-cc101d813220','2025-12-09',88732.06,443660.52),
 ('290','1008 D','67240a63-fd0f-4d21-8dd6-5ea34e4c0c76','2025-11-17',109571.60,349191.04),
 ('307','1707 A','6065c4ff-1bb2-42ab-85f5-b769f7f97253','2025-12-05',87161.76,435808.19),
 ('311','803 A','16f0a4fd-e669-49be-814d-76789ed6bb4c','2025-11-24',82088.56,410443.77),
 ('315','406 A','36965dcf-4d05-42f6-b066-ac480bb1a561','2026-04-09',63308.76,292366.06),
 ('341','503 B','9581a13a-5d34-44a0-b54d-62a841d757a7','2026-03-25',79622.00,398372.68),
 ('354','510 C','908fbcdb-ac7a-45f0-95c1-5e4f45e67736','2026-01-15',75880.20,379402.55),
 ('355','1707 A','83b8f727-e9a9-4a88-846c-272f6970356a','2025-12-04',87161.76,435808.19),
 ('361','1707 A','7787a154-ec62-469f-b330-d9b3b6142dd2','2026-01-29',88913.20,444567.92),
 ('362','906 A','5cbed8bd-36a1-4ec6-ace1-0aa9da2023b1','2026-01-15',175999.86,372000.00),
 ('380','507 A','54497681-10e3-4a1a-89d8-b3356fdec09a','2026-02-08',57370.04,292366.06),
 ('384','604 D','d77a7a35-7c36-450a-b4af-9648bc44adc7','2026-03-17',85379.61,426900.16),
]
dist_vendas = {d[2] for d in distratos}
dist_unid = {}
for d in distratos: dist_unid.setdefault(d[1], []).append(d[0])
b10_by_unit = {c['unidade']: c['numero_contrato'] for c in b10['casos']}

TIPO = {'parcela_entrada':'Parcela mensal','sinal':'Sinal/entrada','balao':'Parcela balão'}
def br_date(s):
    y,m,d = s.split('-'); return f'{d}/{m}/{y}'

# ---------- estilos ----------
FONT = 'Arial'
H_FILL = PatternFill('solid', fgColor='1F4E78')
H_FONT = Font(name=FONT, bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color='1F4E78')
SUB_FONT = Font(name=FONT, size=10, color='444444')
CELL_FONT = Font(name=FONT, size=10)
RESP_FILL = PatternFill('solid', fgColor='FFF2CC')
RESP_HFILL = PatternFill('solid', fgColor='BF8F00')
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
MONEY = '"R$" #,##0.00'

wb = Workbook()

def header_block(ws, title, subtitle, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1,1,title); c.font = TITLE_FONT; c.alignment = Alignment(vertical='center')
    ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c2 = ws.cell(2,1,subtitle); c2.font = SUB_FONT; c2.alignment = WRAP_TOP
    ws.row_dimensions[2].height = 48

def write_table(ws, start_row, headers, rows, widths, resp_cols, money_cols, center_cols):
    r = start_row
    for j,h in enumerate(headers, start=1):
        c = ws.cell(r, j, h)
        c.font = H_FONT; c.alignment = CENTER; c.border = BORDER
        c.fill = RESP_HFILL if j-1 in resp_cols else H_FILL
    ws.row_dimensions[r].height = 32
    for j,w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(1,j).column_letter].width = w
    for row in rows:
        r += 1
        for j,val in enumerate(row, start=1):
            c = ws.cell(r, j, val); c.font = CELL_FONT; c.border = BORDER
            if j-1 in money_cols: c.number_format = MONEY
            if j-1 in center_cols: c.alignment = Alignment(horizontal='center', vertical='center')
            else: c.alignment = WRAP_TOP
            if j-1 in resp_cols: c.fill = RESP_FILL
    ws.freeze_panes = ws.cell(start_row+1, 1)

# ===================== ABA INÍCIO =====================
ws0 = wb.active; ws0.title = 'Início'
ws0.sheet_view.showGridLines = False
ws0.column_dimensions['A'].width = 3
ws0.column_dimensions['B'].width = 100
def line(r, txt, font):
    c = ws0.cell(r,2,txt); c.font = font; c.alignment = WRAP_TOP
line(2,'Conferência de contratos — IM x Sienge', TITLE_FONT)
line(4,'Olá! Levantamos três pontos onde o nosso sistema e o Sienge não estão batendo. '
      'Precisamos da sua conferência no Sienge para corrigir. Cada aba abaixo trata de um assunto. '
      'A coluna amarela (RESPOSTA) é para você preencher.', Font(name=FONT, size=11))
itens = [
 ('Aba "1. Parcelas a conferir"','Parcelas que aparecem no nosso sistema mas não encontramos no Sienge. Precisamos saber se cada uma realmente existe / foi paga pelo cliente.'),
 ('Aba "2. Saldo divergente"','Contratos em que o saldo devedor (pró-soluto) do nosso sistema é diferente da soma das parcelas no Sienge. Precisamos saber qual valor é o correto.'),
 ('Aba "3. Distratos a confirmar"','Contratos que constam como distratados no Sienge mas ainda aparecem como ativos aqui. A comissão já paga é mantida — só precisamos confirmar o distrato.'),
 ('Aba "4. Em análise interna"','Contratos com possível duplicidade de parcelas. Esses ficam com a nossa equipe — são informativos, não precisam de resposta.'),
]
r = 6
for t,d in itens:
    c = ws0.cell(r,2,t); c.font = Font(name=FONT, bold=True, size=11, color='1F4E78'); r+=1
    c = ws0.cell(r,2,'   '+d); c.font = Font(name=FONT, size=10); c.alignment = WRAP_TOP
    ws0.row_dimensions[r].height = 30; r+=2
line(r+1,'Qualquer dúvida sobre o que cada coluna significa, é só chamar. Obrigado!', SUB_FONT)

# ===================== ABA 1 — PARCELAS A CONFERIR =====================
ws1 = wb.create_sheet('1. Parcelas a conferir')
ws1.sheet_view.showGridLines = False
header_block(ws1,
  '1. Parcelas que não encontramos no Sienge',
  'Cada linha é uma parcela registrada no NOSSO sistema que não localizamos no Sienge. '
  'Por favor, confirme no Sienge se a parcela existe e foi paga pelo cliente. Se NÃO existir, vamos removê-la do nosso sistema.',
  9)
rows1 = []
casos1 = [c for c in b9['casos'] if c['grupo'] in (1,2)]
# ordenar: distrato primeiro, depois por contrato
for c in sorted(casos1, key=lambda x:(x['venda_id'] not in dist_vendas, int(x['numero_contrato']))):
    obs = 'Contrato distratado' if c['venda_id'] in dist_vendas else ''
    for l in c['estado_atual']['linhas']:
        rows1.append([
            'c'+c['numero_contrato'], c['unidade'], (c.get('cliente') or '').strip().title(),
            TIPO.get(l['tipo'], l['tipo']),
            float(l['valor']),
            'Paga' if l['status']=='pago' else 'Pendente',
            br_date(l['data_pagamento']) if l.get('data_pagamento') else '—',
            obs, ''
        ])
write_table(ws1, 4,
  ['Contrato','Unidade','Cliente','Tipo de parcela','Valor da parcela','Situação no nosso sistema','Data do pagamento (nosso sistema)','Observação','RESPOSTA: a parcela existe no Sienge? (Sim / Não)'],
  rows1,
  [10,10,30,16,16,16,16,18,28],
  resp_cols=[8], money_cols=[4], center_cols=[0,1,5,6])

# ===================== ABA 2 — SALDO DIVERGENTE =====================
ws2 = wb.create_sheet('2. Saldo divergente')
ws2.sheet_view.showGridLines = False
header_block(ws2,
  '2. Saldo devedor (pró-soluto) diferente do Sienge',
  'Para estes contratos, o saldo devedor registrado aqui é diferente da soma das parcelas no Sienge. '
  'Por favor, indique qual valor é o correto.', 8)
rows2 = []
for c in sorted(b10['casos'], key=lambda x:-abs(x['diferenca'])):
    u = c['unidade']
    obs = ''
    if u in dist_unid: obs = 'Unidade revendida (o contrato anterior foi distratado)'
    rows2.append([
        'c'+c['numero_contrato'], u,
        float(c['pro_soluto_local']), float(c['soma_income_sienge']),
        float(c['diferenca']), c['parcelas_pagas'], obs, ''
    ])
write_table(ws2, 4,
  ['Contrato','Unidade','Saldo no nosso sistema','Saldo no Sienge','Diferença','Parcelas já pagas','Observação','RESPOSTA: qual valor é o correto?'],
  rows2,
  [10,10,18,18,16,14,34,28],
  resp_cols=[7], money_cols=[2,3,4], center_cols=[0,1,5])

# ===================== ABA 3 — DISTRATOS =====================
ws3 = wb.create_sheet('3. Distratos a confirmar')
ws3.sheet_view.showGridLines = False
header_block(ws3,
  '3. Contratos distratados (confirmar)',
  'Estes contratos constam como DISTRATADOS no Sienge mas ainda aparecem como ativos no nosso sistema. '
  'A comissão já paga será mantida. Por favor, confirme o distrato e a data.', 6)
rows3 = []
for contrato,unid,vid,data,pro,valor in sorted(distratos, key=lambda x:int(x[0])):
    obs = []
    if len(dist_unid[unid])>1: obs.append(f'Unidade distratada {len(dist_unid[unid])}x')
    if unid in b10_by_unit: obs.append(f'Revendida (contrato novo c{b10_by_unit[unid]})')
    rows3.append(['c'+contrato, unid, br_date(data), float(valor), '; '.join(obs), ''])
write_table(ws3, 4,
  ['Contrato','Unidade','Data do distrato (Sienge)','Valor da venda','Observação','RESPOSTA: confirma o distrato? (Sim / Não)'],
  rows3,
  [10,10,20,18,34,30],
  resp_cols=[5], money_cols=[3], center_cols=[0,1,2])

# ===================== ABA 4 — EM ANÁLISE INTERNA =====================
ws4 = wb.create_sheet('4. Em análise interna')
ws4.sheet_view.showGridLines = False
header_block(ws4,
  '4. Contratos com possível duplicidade (análise interna)',
  'Estes contratos têm parcelas que podem estar duplicadas no sistema. Ficam com a nossa equipe para '
  'análise individual — são informativos e NÃO precisam de resposta da controladoria.', 4)
rows4 = []
for c in sorted([x for x in b9['casos'] if x['grupo']==3], key=lambda x:int(x['numero_contrato'])):
    rows4.append(['c'+c['numero_contrato'], c['unidade'], (c.get('cliente') or '').strip().title(),
                  'Possível parcela duplicada — verificar'])
write_table(ws4, 4,
  ['Contrato','Unidade','Cliente','Situação'],
  rows4, [10,10,32,40], resp_cols=[], money_cols=[], center_cols=[0,1])

out_dir = os.path.join(ROOT,'docs','controladoria')
os.makedirs(out_dir, exist_ok=True)
out = os.path.join(out_dir,'conferencia-sienge-2026-06-02.xlsx')
wb.save(out)
print('Salvo:', out)
print('Linhas: aba1=%d aba2=%d aba3=%d aba4=%d' % (len(rows1),len(rows2),len(rows3),len(rows4)))
