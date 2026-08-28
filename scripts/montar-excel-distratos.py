# Planilha de DISTRATOS a partir do JSON de scripts/extrair-relatorio-distratos.mjs.
# Aba 1 "Panorama" (os totais + o estado da cura) · Aba 2 "Distratos" (um por linha).
import json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ORIGEM, DESTINO = sys.argv[1], sys.argv[2]
d = json.load(open(ORIGEM, encoding='utf-8'))
T, M = d['totais'], d['meta']

FONTE = 'Arial'
AZUL, VERMELHO, CINZA = '1F3864', 'C00000', 'F2F2F2'
VERDE_TX, BRANCO, CREME, ROSA = '006100', 'FFFFFF', 'FFF2CC', 'FDE9E9'
MOEDA = 'R$ #,##0.00;[Red](R$ #,##0.00);-'
thin = Side(style='thin', color='BFBFBF')
BORDA = Border(left=thin, right=thin, top=thin, bottom=thin)

def data_br(iso):
    if not iso:
        return ''
    a, m, dia = str(iso)[:10].split('-')
    return f'{dia}/{m}/{a}'

def titulo(ws, row, texto, ncols, tam=13, cor=AZUL):
    c = ws.cell(row=row, column=1, value=texto)
    c.font = Font(name=FONTE, size=tam, bold=True, color=BRANCO)
    c.fill = PatternFill('solid', fgColor=cor)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.row_dimensions[row].height = 22

def cabecalho(ws, row, headers, destaques=()):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name=FONTE, size=10, bold=True, color=BRANCO)
        c.fill = PatternFill('solid', fgColor=VERMELHO if i in destaques else AZUL)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDA
    ws.row_dimensions[row].height = 34

wb = Workbook()

# ═══════════════════════════ Panorama ═══════════════════════════
ws = wb.active
ws.title = 'Panorama'
NC = 4
titulo(ws, 1, 'DISTRATOS - PANORAMA', NC, 14)
sub = ws.cell(row=2, column=1, value=(
    'Todo contrato distratado da base, com o que o cliente REALMENTE pagou antes de cancelar '
    'e o que foi baixa-em-massa (efeito contabil do cancelamento, nao dinheiro). '
    'Regra: vale como pago a parcela cujo VENCIMENTO e anterior ou igual a data do distrato — '
    'usar a data de pagamento como regua falha, porque a baixa em massa reescreve a data das '
    'parcelas que ja estavam legitimamente pagas.'))
sub.font = Font(name=FONTE, size=10, italic=True, color='595959')
sub.alignment = Alignment(wrap_text=True, vertical='top')
ws.merge_cells('A2:D2')
ws.row_dimensions[2].height = 58
ws.cell(row=3, column=1, value=f'Gerado em {M["geradoEm"][:10]}  ·  regenerar: node scripts/extrair-relatorio-distratos.mjs').font = Font(name=FONTE, size=9, color='808080')

blocos = [
    ('O QUE FOI CANCELADO', [
        ('Contratos distratados', T['distratos'], '#,##0'),
        ('Corretores afetados', T['corretores_afetados'], '#,##0'),
        ('VGV distratado (nao e mais carteira)', T['vgv_distratado'], MOEDA),
    ]),
    ('O QUE O CLIENTE PAGOU DE VERDADE (antes do distrato)', [
        ('Valor recebido dos clientes', T['valor_pago_real'], MOEDA),
        ('Comissao total gerada (todos os cargos)', T['comissao_total_real'], MOEDA),
        ('Fatia do corretor', T['fatia_corretor_real'], MOEDA),
    ]),
    ('BAIXA-EM-MASSA (o que o Sienge marcou como pago e nunca foi dinheiro)', [
        ('Parcelas ja canceladas pela cura', T['parcelas_curadas'], '#,##0'),
        ('Comissao falsa removida', T['comissao_falsa_removida'], MOEDA),
        ('Parcelas AINDA marcadas como pagas (pendente de cura)', T['parcelas_falsas_vivas'], '#,##0'),
        ('Comissao falsa ainda no sistema', T['comissao_falsa_viva'], MOEDA),
    ]),
]
r = 5
for nome, itens in blocos:
    alerta = nome.startswith('BAIXA') and T['parcelas_falsas_vivas'] > 0
    titulo(ws, r, nome, NC, 11, cor=VERMELHO if alerta else AZUL)
    r += 1
    for rot, val, fmt in itens:
        ws.cell(row=r, column=1, value=rot).font = Font(name=FONTE, size=10)
        c = ws.cell(row=r, column=3, value=val)
        c.number_format = fmt
        ruim = 'AINDA' in rot or 'ainda' in rot
        c.font = Font(name=FONTE, size=11, bold=True,
                      color=VERMELHO if (ruim and val) else AZUL)
        if ruim and val:
            c.fill = PatternFill('solid', fgColor=ROSA)
        for i in range(1, NC + 1):
            ws.cell(row=r, column=i).border = BORDA
        r += 1
    r += 1

st = ws.cell(row=r, column=1, value=(
    'SITUACAO: nenhuma baixa-em-massa pendente — todos os distratos estao curados.'
    if T['parcelas_falsas_vivas'] == 0 else
    f'ATENCAO: {T["parcelas_falsas_vivas"]} parcelas de distrato ainda contam como pagas. '
    f'Rode: node scripts/curar-distrato-apply.mjs --apply'))
st.font = Font(name=FONTE, size=11, bold=True,
               color=VERDE_TX if T['parcelas_falsas_vivas'] == 0 else VERMELHO)
st.fill = PatternFill('solid', fgColor=CREME if T['parcelas_falsas_vivas'] == 0 else ROSA)
st.alignment = Alignment(wrap_text=True, vertical='center')
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
ws.row_dimensions[r].height = 30

r += 2
nota = ws.cell(row=r, column=1, value=(
    'Desde 28/08/2026 a limpeza roda sozinha todo dia no cron, depois da sincronizacao com o '
    'Sienge. Antes era manual, e por isso um distrato de 25/08 (apto 1002 D) deixou R$ 12.868,71 '
    'de comissao inexistente no painel de um corretor por tres dias — foi o que a controladoria '
    'flagrou. Contrato distratado nao aparece nas telas do corretor; aparece aqui e no relatorio '
    'do admin, que e quem audita.'))
nota.font = Font(name=FONTE, size=9, italic=True, color='595959')
nota.alignment = Alignment(wrap_text=True, vertical='top')
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
ws.row_dimensions[r].height = 60

for col, w in zip('ABCD', [52, 4, 24, 12]):
    ws.column_dimensions[col].width = w

# ═══════════════════════════ Distratos ═══════════════════════════
s = wb.create_sheet('Distratos')
COLS = [
    ('Unidade', 'unidade', None, 12),
    ('Cliente', 'cliente', None, 30),
    ('Corretor', 'corretor', None, 28),
    ('Contrato', 'contrato', None, 10),
    ('Data da venda', 'data_venda', 'data', 13),
    ('Data do distrato', 'data_distrato', 'data', 14),
    ('Valor da venda', 'valor_venda', MOEDA, 16),
    ('Pro-soluto', 'valor_pro_soluto', MOEDA, 14),
    ('Parcelas pagas (reais)', 'parcelas_pagas_reais', '#,##0', 11),
    ('Valor recebido do cliente', 'valor_pago_real', MOEDA, 17),
    ('Comissao total gerada', 'comissao_total_real', MOEDA, 16),
    ('Fatia do corretor', 'fatia_corretor_real', MOEDA, 14),
    ('1o pagamento', 'primeiro_pagamento', 'data', 12),
    ('Ultimo pagamento', 'ultimo_pagamento', 'data', 12),
    ('Parcelas canceladas pela cura', 'parcelas_curadas', '#,##0', 12),
    ('Comissao falsa removida', 'comissao_falsa_removida', MOEDA, 16),
    ('AINDA marcadas pagas', 'parcelas_falsas_vivas', '#,##0', 11),
    ('Pendencia', 'pendencia', None, 40),
]
titulo(s, 1, f'DISTRATOS ({T["distratos"]}) - do mais recente ao mais antigo', len(COLS), 13)
cabecalho(s, 2, [c[0] for c in COLS], destaques=(17, 18))
r = 3
ini = r
for l in d['distratos']:
    for i, (_, chave, fmt, _w) in enumerate(COLS, start=1):
        v = l.get(chave, '')
        c = s.cell(row=r, column=i, value=data_br(v) if fmt == 'data' else v)
        if fmt and fmt != 'data':
            c.number_format = fmt
        c.border = BORDA
        c.font = Font(name=FONTE, size=10)
    s.cell(row=r, column=1).font = Font(name=FONTE, size=10, bold=True)
    if l.get('parcelas_falsas_vivas'):
        for i in range(1, len(COLS) + 1):
            s.cell(row=r, column=i).fill = PatternFill('solid', fgColor=ROSA)
        s.cell(row=r, column=17).font = Font(name=FONTE, size=10, bold=True, color=VERMELHO)
    r += 1

s.cell(row=r, column=1, value='TOTAL').font = Font(name=FONTE, size=11, bold=True)
for i, (_, chave, fmt, _w) in enumerate(COLS, start=1):
    if fmt in (MOEDA, '#,##0'):
        L = get_column_letter(i)
        c = s.cell(row=r, column=i, value=f'=SUM({L}{ini}:{L}{r - 1})')
        c.number_format = fmt
        c.font = Font(name=FONTE, size=11, bold=True)
for i in range(1, len(COLS) + 1):
    s.cell(row=r, column=i).fill = PatternFill('solid', fgColor=CINZA)
    s.cell(row=r, column=i).border = BORDA

for i, (_, _, _, w) in enumerate(COLS, start=1):
    s.column_dimensions[get_column_letter(i)].width = w
s.freeze_panes = 'C3'
s.auto_filter.ref = f'A2:{get_column_letter(len(COLS))}{r - 1}'

wb.save(DESTINO)
print(f'ok: {DESTINO} — Panorama + {T["distratos"]} distratos')
